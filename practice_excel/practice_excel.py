#create a new file using python

import xlsxwriter
import openpyxl as op

# data9 = xlsxwriter.Workbook("data9.xlsx") #creating a workbook using the built in method and assigning a name to the file
# worksheet=data9.add_worksheet("data9.xlsx") #creating an object
#
#
# #note - rows and columns are zero indexed
# #the first cell in a worksheet A1 is (0,0), B1 is (0,1)
# worksheet.write("A1","data 9")
#
# row=1
# column = 0
# content = ["Jonny","GeeGee","Jonny","Ben"]
# for name in content:
#     worksheet.write(row,column,name)
#     row +=1
#
# data9.close() #this saves and closes the file



#Fetch the data from the excel file using openpyxl
# path = "data9.xlsx"# give location of the file
# data9_object = op.load_workbook(path)#
# data9_object = data9_object.active
# data9_sheet = data9_object.title
#
# get_data = data9_object.cell(row=1,column=1)
# print(get_data.value)
#
# # print(data9_object.max_column,data9_object.max_row)
# # print(data9_sheet)
# # data9_object.title = "Trainees"
# # data9_sheet = data9_object.title
# # print(data9_sheet)
# #
# # data9_object.cell(column=1, row=1).value = "Data10"
# # data9_object.cell(column=1, row=2).value = "Todger"
# # print(data9_object.cell(column=1, row=2).value)
# #
# # data9_object.save("data9.xlsx")
#
#
# #Write a program in python to copy data from data9 file to a new file called 'data10'
# path = "data9.xlsx"# give location of the file
# data9_object = op.load_workbook(path)#
# data9_ws = data9_object.worksheets[0] #this gets the data from the file
#
#
# data10_file = xlsxwriter.Workbook("data10.xlsx") #creating a workbook using the built in method and assigning a name to the file
# sheet=data10_file.add_worksheet("data10.xlsx")
# data10_file.close()
#
# #call the new workbook back in
# dat10 = op.load_workbook("data10.xlsx")
# data10_ws =dat10.worksheets[0]
#
# mr = data9_ws.max_row
# mc = data9_ws.max_column
#
# # copying the cell values from source
# # excel file to destination excel file
# for i in range(1, mr + 1):
#     for j in range(1, mc + 1):
#         # reading cell value from source excel file
#         c = data9_ws.cell(row=i, column=j)
#
#         # writing the read value to destination excel file
#         data10_ws.cell(row=i, column=j).value = c.value
#
#     # saving the destination excel file
# dat10.save(str("data10.xlsx"))