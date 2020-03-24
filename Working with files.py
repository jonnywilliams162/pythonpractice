###################################################
#### WORKING WITH, READING IN AND SAVING FILES ####
###################################################

#'open()' takes two parameters: 'filename and mode'
#four different modes:
#'r' - read (default value)
#'a' - append
#'w' - write
#'x' - creates the specified file
#should specify if the file should be handled as binary or text mode 't' or 'b'

#syntax:
# file = open("demofile.txt")
# file = open("demofile.txt","rt") #these are the same as r for read and t for text are the defau;t
# print(file.read()) #will print out the contents of the file
# print(file.read(5))#will display only first 5 characters of the file
# print(file.readline())#will return one line, by calling this twice you can see the first two lines
#
# for x in file
#     print(x) #this loop will print the whole file line by line
#
# #it is good practice to close the file once done with it to save space and memory
# file.close()

###writing to an existing file ###
# f = open("demofile.txt", "a") #the "a" allows you to append to the existing file
# f.write("more content for the file")
# f.close()
#
# f=open("demofile.txt","w") #"w" will overwrite the existing content of the file!
# f.write("woops i have deleted the content")
# f.close

### Create a new file ###
#can use the modes: 'a,x or w'
#'x' is the proper way, the others will just make one if there isnt one with the name you put in already

# f=open("myfile.txt","x")
# f.write("first bit of content for myfile")

### Delete files ###
# must import the 'os' module and run its 'os.remove()' function
import os
#os.remove("demofile.txt")

#check if file exists
# if os.path.exists("demofile.txt"):
#     os.remove("demofile.txt")
# else:
#     print("the file does not exist")

#to delete an entire folder use 'os.rmdir()' remove directory
# os.rmdir("myfolder")



###################
### EXCEL FILES ###
###################
#openpyxl and xlsxwriter libraries
#using these libraries to write and read to/from excel files
import openpyxl as op
import xlsxwriter as xls

#create a python folder and a file inside
# data9 = xls.Workbook("data9.xlsx") #creating a workbook using the built in method and assigning a name to the file
# worksheet=data9.add_worksheet("data9.xlsx") #creating an object
# #note - rows and columns are zero indexed
# #the first cell in a worksheet A1 is (0,0), B1 is (0,1)
# worksheet.write("A1","data 9")
#
# row=1
# column = 0
# content = ["Jonny","GeeGee","Jonny","Ben"]
# for name in content:
#     worksheet.write(row,column,name)
#     row +=1
# data9.close() #this saves and closes the file

#Fetch the data from the excel file using openpyxl
# path = "practice_excel/data9.xlsx"# give location of the file
# data9_object = op.load_workbook(path)#
# data9_object = data9_object.active
#
# get_data = data9_object.cell(row=1,column=1)
# print(get_data.value)

# print(data9_object.max_column,data9_object.max_row)
# # print(data9_sheet)
# # data9_object.title = "Trainees"
# # data9_sheet = data9_object.title
# # print(data9_sheet)
# #
# # data9_object.cell(column=1, row=1).value = "Data10"
# # data9_object.cell(column=1, row=2).value = "Todger"
# # print(data9_object.cell(column=1, row=2).value)
# #
# # data9_object.save("data9.xlsx")
#
#
# #Write a program in python to copy data from data9 file to a new file called 'data10'
# path = "data9.xlsx"# give location of the file
# data9_object = op.load_workbook(path)#
# data9_ws = data9_object.worksheets[0] #this gets the data from the file
#
#
# data10_file = xlsxwriter.Workbook("data10.xlsx") #creating a workbook using the built in method and assigning a name to the file
# sheet=data10_file.add_worksheet("data10.xlsx")
# data10_file.close()
#
# #call the new workbook back in
# dat10 = op.load_workbook("data10.xlsx")
# data10_ws =dat10.worksheets[0]
#
# mr = data9_ws.max_row
# mc = data9_ws.max_column
#
# # copying the cell values from source
# # excel file to destination excel file
# for i in range(1, mr + 1):
#     for j in range(1, mc + 1):
#         # reading cell value from source excel file
#         c = data9_ws.cell(row=i, column=j)
#
#         # writing the read value to destination excel file
#         data10_ws.cell(row=i, column=j).value = c.value
#
#     # saving the destination excel file
# dat10.save(str("data10.xlsx"))



#################
### CSV FILES ###
#################

